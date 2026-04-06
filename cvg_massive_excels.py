#!/usr/bin/env python3
"""
Carga masiva genérica Excel -> PostgreSQL basada en metadata de tabla + config.ini.

Objetivo:
- Reutilizable para cualquier Excel/tabla destino.
- Escalable y mantenible por funciones.
- Validación por tipos reales de PostgreSQL.
"""

from __future__ import annotations

import argparse
from datetime import date, datetime, timedelta
import configparser
import difflib
import json
import math
import re
import shutil
import sys
import unicodedata
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Dict, List

import pandas as pd
import psycopg2
from openpyxl import load_workbook
from psycopg2 import sql
from psycopg2.extras import execute_values


SPANISH_MONTHS = {
    "ene": 1,
    "feb": 2,
    "mar": 3,
    "abr": 4,
    "may": 5,
    "jun": 6,
    "jul": 7,
    "ago": 8,
    "sep": 9,
    "oct": 10,
    "nov": 11,
    "dic": 12,
}

EARLY_STOP_SCORE = 200
HEADER_MIN_SCORE = 10
IGNORE_COLUMN = "__IGNORAR__"
DATE_TYPES = {"date"}
TIMESTAMP_TYPES = {"timestamp without time zone", "timestamp with time zone"}
NUMERIC_TYPES = {
    "smallint",
    "integer",
    "bigint",
    "numeric",
    "decimal",
    "real",
    "double precision",
}
BOOL_TYPES = {"boolean"}


@dataclass
class ColumnMeta:
    name: str
    data_type: str
    is_nullable: bool
    column_default: str | None
    is_identity: bool
    is_generated: bool


@dataclass
class ValidationResult:
    valid_df: pd.DataFrame
    invalid_df: pd.DataFrame
    error_messages: pd.Series


class UserCancelledLoad(Exception):
    """Interrupción controlada por usuario; no debe tratarse como error."""
    pass


# =========================================================
# LOGGING
# =========================================================

def setup_logging(log_file: Path | None) -> None:
    if not log_file:
        return

    log_file.parent.mkdir(parents=True, exist_ok=True)

    class Tee:
        def __init__(self, *streams):
            self.streams = streams

        def write(self, data):
            for s in self.streams:
                s.write(data)
                s.flush()
            return len(data)

        def flush(self):
            for s in self.streams:
                s.flush()

    fh = log_file.open("a", encoding="utf-8")
    ts = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    fh.write(f"\n===== RUN START {ts} =====\n")
    fh.flush()

    sys.stdout = Tee(sys.stdout, fh)
    sys.stderr = Tee(sys.stderr, fh)


# =========================================================
# HELPERS
# =========================================================


def save_missing_column_suggestions_to_mapping(
    mapping_path: Path,
    table_section: str,
    suggestions_df: pd.DataFrame,
) -> None:
    cp = load_mapping_store(mapping_path)

    if table_section not in cp:
        cp[table_section] = {}

    useful_df = suggestions_df[suggestions_df["sugerencia_util"]].copy()

    for _, row in useful_df.iterrows():
        excel_col = str(row["excel_columna_ignorada"]).strip()
        table_col = str(row["tabla_columna_faltante"]).strip()

        if not excel_col or not table_col:
            continue

        cp[table_section][excel_col] = table_col

    with mapping_path.open("w", encoding="utf-8") as f:
        cp.write(f)
        

def resolve_path(base_dir: Path, value: str, default: str | None = None) -> Path:
    raw = value if value else (default or "")
    p = Path(raw)
    return p if p.is_absolute() else (base_dir / p).resolve()


def canonicalize_header(header: str) -> str:
    txt = str(header).strip().lower()
    txt = re.sub(r"\([^\)]*\)", " ", txt)
    txt = unicodedata.normalize("NFKD", txt)
    txt = "".join(ch for ch in txt if not unicodedata.combining(ch))
    txt = re.sub(r"[^a-z0-9]+", " ", txt)
    txt = re.sub(r"\s+", " ", txt).strip()
    return txt.replace(" ", "")


def to_snake_name(header: str) -> str:
    txt = str(header).strip().lower()
    txt = re.sub(r"\([^\)]*\)", " ", txt)
    txt = unicodedata.normalize("NFKD", txt)
    txt = "".join(ch for ch in txt if not unicodedata.combining(ch))
    txt = re.sub(r"[^a-z0-9]+", "_", txt)
    txt = re.sub(r"_+", "_", txt).strip("_")
    return txt


def normalize_table_identifier(table_schema: str, table_name: str) -> str:
    return f"{table_schema}.{table_name}"


def ask_yes_no(question: str) -> bool:
    while True:
        ans = input(f"{question} [s/n] (s=si, n=no): ").strip().lower()
        if ans in {"si", "s", "yes", "y"}:
            return True
        if ans in {"no", "n"}:
            return False
        print("Respuesta no válida. Usa: s (si) / n (no)")


def prompt_choice(title: str, options: List[str]) -> str:
    if not options:
        raise ValueError(f"No hay opciones disponibles para: {title}")

    print(f"\n[SELECCION] {title}")
    for i, opt in enumerate(options, 1):
        print(f"  {i}. {opt}")

    while True:
        raw = input(f"Selecciona una opción (1-{len(options)}): ").strip()
        if raw.isdigit():
            idx = int(raw)
            if 1 <= idx <= len(options):
                return options[idx - 1]
        print("Selección inválida, intenta de nuevo.")


def choose_load_mode(input_dir: Path, retry_input_dir: Path) -> Path:
    """Pregunta si la carga es inicial o reintento y devuelve carpeta fuente."""
    print("\n[SELECCION] Tipo de carga")
    print("  1. Primer insert (usa input_dir)")
    print("  2. Reintento de inserts (usa retry_input_dir)")

    while True:
        raw = input("Selecciona una opción (1-2): ").strip()
        if raw == "1":
            return input_dir
        if raw == "2":
            return retry_input_dir
        print("Selección inválida, intenta de nuevo.")


def safe_delete(path: Path | None) -> None:
    if not path:
        return
    try:
        if path.exists() and path.is_file():
            path.unlink()
    except Exception:
        pass


def safe_str_cell(value: Any) -> str:
    if value is None or pd.isna(value):
        return ""
    return str(value).strip()


# =========================================================
# CONFIG
# =========================================================

def load_config(config_path: Path) -> configparser.ConfigParser:
    cfg = configparser.ConfigParser()
    cfg.optionxform = str
    if not config_path.exists():
        raise FileNotFoundError(f"No existe config.ini: {config_path}")
    cfg.read(config_path, encoding="utf-8")
    for section in ["postgres", "target", "input", "output", "run"]:
        if section not in cfg:
            raise KeyError(f"Falta sección [{section}] en {config_path}")
    return cfg


def get_db_params(cfg: configparser.ConfigParser) -> Dict[str, str]:
    return dict(cfg["postgres"])


# =========================================================
# DB METADATA
# =========================================================

def fetch_available_schemas(conn_params: Dict[str, str]) -> List[str]:
    query = """
        SELECT schema_name
        FROM information_schema.schemata
        WHERE schema_name NOT IN ('information_schema', 'pg_catalog', 'pg_toast')
          AND schema_name NOT LIKE 'pg_temp_%'
          AND schema_name NOT LIKE 'pg_toast_temp_%'
        ORDER BY schema_name
    """
    with psycopg2.connect(**conn_params) as conn:
        with conn.cursor() as cur:
            cur.execute(query)
            return [r[0] for r in cur.fetchall()]


def fetch_tables_in_schema(conn_params: Dict[str, str], schema: str) -> List[str]:
    query = """
        SELECT table_name
        FROM information_schema.tables
        WHERE table_schema = %s
          AND table_type = 'BASE TABLE'
        ORDER BY table_name
    """
    with psycopg2.connect(**conn_params) as conn:
        with conn.cursor() as cur:
            cur.execute(query, (schema,))
            return [r[0] for r in cur.fetchall()]


def schema_exists(conn_params: Dict[str, str], schema: str) -> bool:
    query = """
        SELECT 1
        FROM information_schema.schemata
        WHERE schema_name = %s
        LIMIT 1
    """
    with psycopg2.connect(**conn_params) as conn:
        with conn.cursor() as cur:
            cur.execute(query, (schema,))
            return cur.fetchone() is not None


def table_exists(conn_params: Dict[str, str], schema: str, table: str) -> bool:
    query = """
        SELECT 1
        FROM information_schema.tables
        WHERE table_schema = %s
          AND table_name = %s
          AND table_type = 'BASE TABLE'
        LIMIT 1
    """
    with psycopg2.connect(**conn_params) as conn:
        with conn.cursor() as cur:
            cur.execute(query, (schema, table))
            return cur.fetchone() is not None


def get_table_metadata(conn_params: Dict[str, str], schema: str, table: str) -> List[ColumnMeta]:
    query = """
        SELECT
            column_name,
            data_type,
            is_nullable,
            column_default,
            is_identity,
            is_generated
        FROM information_schema.columns
        WHERE table_schema = %s
          AND table_name = %s
        ORDER BY ordinal_position
    """
    with psycopg2.connect(**conn_params) as conn:
        with conn.cursor() as cur:
            cur.execute(query, (schema, table))
            rows = cur.fetchall()

    if not rows:
        raise ValueError(
            f"No se encontró metadata para {schema}.{table}. "
            "Verifica [target], [target_defensa_options], el dbname/usuario y permisos del usuario. "
            "También puedes ejecutar con --interactive-target para elegir schema/tabla desde consola."
        )

    return [
        ColumnMeta(
            name=r[0],
            data_type=r[1],
            is_nullable=(r[2] == "YES"),
            column_default=r[3],
            is_identity=(r[4] == "YES"),
            is_generated=(r[5] != "NEVER"),
        )
        for r in rows
    ]


def get_insertable_columns(metadata: List[ColumnMeta]) -> List[str]:
    cols = []
    for c in metadata:
        if c.is_identity or c.is_generated:
            continue
        if c.column_default and "nextval(" in c.column_default:
            continue
        cols.append(c.name)
    return cols


# =========================================================
# INPUT / EXCEL
# =========================================================

def pick_excel_file(input_dir: Path, file_name: str | None = None) -> Path:
    if file_name:
        file_path = input_dir / file_name
        if not file_path.exists():
            raise FileNotFoundError(f"No existe el fichero: {file_path}")
        return file_path

    candidates = sorted(
        list(input_dir.glob("*.xlsx")) + list(input_dir.glob("*.xls")),
        key=lambda p: p.stat().st_mtime,
        reverse=True,
    )
    if not candidates:
        raise FileNotFoundError(f"No hay ficheros Excel en {input_dir}")
    return candidates[0]


def choose_sheet_name(excel_path: Path, preferred_sheet: str | None) -> str:
    xls = pd.ExcelFile(excel_path)
    try:
        sheets = xls.sheet_names
    finally:
        xls.close()

    if not sheets:
        raise ValueError(f"\n>> [ERROR] El Excel {excel_path.name} no contiene hojas.")

    preferred = (preferred_sheet or "").strip()
    if preferred and preferred in sheets:
        return preferred

    if len(sheets) == 1:
        chosen = sheets[0]
        print(f"\n>> [INFO] El Excel tiene una sola hoja. Usando: {chosen}")
        return chosen

    if preferred:
        print(f"\n>> [INFO] La hoja configurada '{preferred}' no existe en {excel_path.name}.")

    print(f">> [INFO] Hojas disponibles: {', '.join(sheets)}")
    return prompt_choice("Selecciona la hoja con datos a cargar", sheets)


def score_header_candidate_row(
    values: List[Any],
    target_columns: List[str],
    similarity_threshold: float,
) -> float:
    cleaned = [safe_str_cell(v) for v in values]
    non_empty = [v for v in cleaned if v != ""]
    if not non_empty:
        return -1.0

    target_keys = [canonicalize_header(c) for c in target_columns]
    target_key_set = set(target_keys)

    score = 0.0
    seen_norm = set()

    for cell in non_empty:
        norm = canonicalize_header(cell)
        if not norm:
            continue

        if norm in seen_norm:
            score -= 0.25
        else:
            seen_norm.add(norm)
            score += 0.25

        if norm in target_key_set:
            score += 6.0
            continue

        best_ratio = 0.0
        for tgt in target_keys:
            ratio = difflib.SequenceMatcher(None, norm, tgt).ratio()
            if ratio > best_ratio:
                best_ratio = ratio

        if best_ratio >= similarity_threshold:
            score += 3.0 + best_ratio
        else:
            if re.search(r"[a-zA-Z]", cell):
                score += 0.4
            if len(cell) <= 60:
                score += 0.2

        if re.fullmatch(r"unnamed[:\s_0-9\-]*", cell.strip().lower()):
            score -= 2.0

    score += min(len(non_empty), 20) * 0.15
    return score


def detect_header_row(
    excel_path: Path,
    sheet_name: str,
    target_columns: List[str],
    similarity_threshold: float,
    scan_rows: int = 25,
) -> int:
    preview = pd.read_excel(excel_path, sheet_name=sheet_name, header=None, nrows=scan_rows)

    if preview.empty:
        return 0

    best_idx = 0
    best_score = float("-inf")
    diagnostics = []

    for idx in range(len(preview)):
        row_values = preview.iloc[idx].tolist()
        score = score_header_candidate_row(
            row_values,
            target_columns=target_columns,
            similarity_threshold=similarity_threshold,
        )
        diagnostics.append((idx, score, [safe_str_cell(v) for v in row_values[:8]]))
        if score > best_score:
            best_score = score
            best_idx = idx

            if score >= EARLY_STOP_SCORE:
                break

    print("\n>> [HEADER] Detección automática de fila de encabezados")

    # guardar detalle solo en log
    if hasattr(sys.stdout, "streams") and len(getattr(sys.stdout, "streams", [])) > 1:
        log_stream = sys.stdout.streams[1]
        for idx, score, sample in diagnostics[:10]:
            sample_txt = " | ".join([s for s in sample if s][:5]) or "(fila vacía)"
            log_stream.write(
                f"  - fila_excel={idx + 1} score={round(score, 4)} sample={sample_txt}\n"
            )
            log_stream.flush()

    print(f">> [HEADER] Fila detectada como header: {best_idx + 1}")
    return best_idx


def resolve_header_row(
    cfg: configparser.ConfigParser,
    excel_path: Path,
    sheet_name: str,
    target_columns: List[str],
    similarity_threshold: float,
) -> int:
    raw_header_row = cfg["input"].get("header_row", fallback="").strip()

    if raw_header_row:
        if raw_header_row.isdigit():
            header_excel_row = int(raw_header_row)
            if header_excel_row < 1:
                raise ValueError("[input].header_row debe ser >= 1")
            print(f"[HEADER] Usando header_row configurado en config.ini: fila {header_excel_row}")
            return header_excel_row - 1
        raise ValueError("[input].header_row debe ser un entero >= 1 (fila Excel)")

    return detect_header_row(
        excel_path=excel_path,
        sheet_name=sheet_name,
        target_columns=target_columns,
        similarity_threshold=similarity_threshold,
    )


def choose_header_mode_interactive(
    cfg: configparser.ConfigParser,
    config_path: Path,
    mapping_path: Path,
    excel_path: Path,
    sheet_name: str,
    target_columns: List[str],
    similarity_threshold: float,
    non_interactive: bool,
) -> int:
    """
    Permite al usuario elegir:
    1) usar header configurado
    2) indicar un nuevo header manual y guardarlo en config.ini
    3) usar detección automática
    Devuelve índice base 0 para pandas.
    """
    table_section = f"{cfg['target'].get('schema')}.{cfg['target'].get('table')}"
    saved_header = get_saved_header(mapping_path, table_section)
    
    if saved_header and non_interactive:
        return saved_header - 1
    
    if non_interactive:
        return resolve_header_row(
            cfg=cfg,
            excel_path=excel_path,
            sheet_name=sheet_name,
            target_columns=target_columns,
            similarity_threshold=similarity_threshold,
        )

    raw_header_row = cfg["input"].get("header_row", fallback="").strip()
    configured_header_excel_row = None

    if raw_header_row:
        if raw_header_row.isdigit() and int(raw_header_row) >= 1:
            configured_header_excel_row = int(raw_header_row)
        else:
            raise ValueError("[input].header_row debe ser un entero >= 1 (fila Excel)")

    print("\n[HEADER] Selección de encabezado")
    print("  1. Usar detección automática")

    if configured_header_excel_row is not None:
        print(f"  2. Usar header configurado actual (fila {configured_header_excel_row})")
        print("  3. Indicar un nuevo header manual y guardarlo en config.ini")
        valid_options = {"1", "2", "3"}
    else:
        print("  2. Indicar un nuevo header manual y guardarlo en config.ini")
        valid_options = {"1", "2"}

    while True:
        choice = input(f"Selecciona una opción ({'/'.join(sorted(valid_options))}): ").strip()

        if choice == "1":
            return detect_header_row(
                excel_path=excel_path,
                sheet_name=sheet_name,
                target_columns=target_columns,
                similarity_threshold=similarity_threshold,
            )

        if configured_header_excel_row is not None and choice == "2":
            print(f"[HEADER] Usando header configurado actual: fila {configured_header_excel_row}")
            return configured_header_excel_row - 1

        if (configured_header_excel_row is not None and choice == "3") or (
            configured_header_excel_row is None and choice == "2"
        ):
            while True:
                raw = input("Indica la fila Excel del header (entero >= 1): ").strip()
                if raw.isdigit() and int(raw) >= 1:
                    new_header_excel_row = int(raw)
                    cfg["input"]["header_row"] = str(new_header_excel_row)
                    with config_path.open("w", encoding="utf-8") as f:
                        cfg.write(f)
                    print(f"[CONFIG] header_row actualizado y guardado en config.ini: fila {new_header_excel_row}")
                    return new_header_excel_row - 1

                print("Valor inválido. Debe ser un entero >= 1.")

        print("Selección inválida, intenta de nuevo.")


def read_excel_with_sheet(excel_path: Path, sheet_name: str, header_row_idx: int) -> pd.DataFrame:
    try:
        return pd.read_excel(excel_path, sheet_name=sheet_name, header=header_row_idx)
    except ValueError as e:
        if "Worksheet named" in str(e):
            xls = pd.ExcelFile(excel_path)
            try:
                available = ", ".join(xls.sheet_names)
            finally:
                xls.close()
            raise ValueError(
                f"La hoja '{sheet_name}' no existe en {excel_path.name}. "
                f"Hojas disponibles: {available}. "
                "Corrige [input].sheet_name en config.ini y vuelve a ejecutar."
            ) from e
        raise


def drop_control_columns(df: pd.DataFrame) -> pd.DataFrame:
    control_cols = {"errores"}
    drop_cols = [c for c in df.columns if str(c).strip().lower() in control_cols]
    if not drop_cols:
        return df
    return df.drop(columns=drop_cols)


def clean_text_values(df: pd.DataFrame) -> pd.DataFrame:
    df = df.copy()
    object_cols = df.select_dtypes(include=["object", "string"]).columns
    replacements = {
        "": pd.NA,
        "-": pd.NA,
        "--": pd.NA,
        "—": pd.NA,
        "–": pd.NA,
        "nan": pd.NA,
        "None": pd.NA,
        "N/A": pd.NA,
        "n/a": pd.NA,
    }
    for col in object_cols:
        df[col] = df[col].astype("string").str.strip().replace(replacements)
    return df


def drop_fully_empty_rows(df: pd.DataFrame, cols: List[str]) -> pd.DataFrame:
    if not cols:
        return df
    mask = df[cols].notna().any(axis=1)
    return df.loc[mask].copy()


def save_header_meta(mapping_path: Path, table_section: str, header_row_excel: int):
    cp = load_mapping_store(mapping_path)
    meta_section = f"{table_section}.__meta__"

    if meta_section not in cp:
        cp[meta_section] = {}

    cp[meta_section]["header_row"] = str(header_row_excel)

    with mapping_path.open("w", encoding="utf-8") as f:
        cp.write(f)


def get_saved_header(mapping_path: Path, table_section: str) -> int | None:
    cp = load_mapping_store(mapping_path)
    meta_section = f"{table_section}.__meta__"

    if meta_section in cp and "header_row" in cp[meta_section]:
        val = cp[meta_section]["header_row"]
        if val.isdigit():
            return int(val)

    return None


def is_header_valid_fast(
    excel_path: Path,
    sheet_name: str,
    header_idx: int,
    target_columns: List[str],
    similarity_threshold: float,
) -> bool:
    try:
        preview = pd.read_excel(excel_path, sheet_name=sheet_name, header=None, nrows=header_idx + 1)
        row = preview.iloc[header_idx].tolist()
        score = score_header_candidate_row(row, target_columns, similarity_threshold)
        return score >= HEADER_MIN_SCORE
    except Exception:
        return False


# =========================================================
# MAPPING
# =========================================================

def get_config_column_map(cfg: configparser.ConfigParser) -> Dict[str, str]:
    custom_map: Dict[str, str] = {}
    if "column_map" in cfg:
        for raw_name, target_col in cfg["column_map"].items():
            custom_map[canonicalize_header(raw_name)] = target_col.strip()
    return custom_map


def get_semidup_section_name(table_section: str) -> str:
    return f"{table_section}.__semidup__"


def load_semiduplicate_approvals(mapping_path: Path, table_section: str) -> Dict[str, str]:
    cp = load_mapping_store(mapping_path)
    section = get_semidup_section_name(table_section)

    if section not in cp:
        return {}

    return {str(k): str(v) for k, v in cp[section].items()}


def save_semiduplicate_approvals(mapping_path: Path, table_section: str, warnings_df: pd.DataFrame) -> None:
    cp = load_mapping_store(mapping_path)
    section = get_semidup_section_name(table_section)

    if section in cp:
        cp.remove_section(section)

    cp[section] = {}

    for _, row in warnings_df.iterrows():
        cp[section][str(row["excel_columna"])] = str(row["tabla_columna_propuesta"])

    with mapping_path.open("w", encoding="utf-8") as f:
        cp.write(f)


def semiduplicate_warnings_changed(mapping_path: Path, table_section: str, warnings_df: pd.DataFrame) -> bool:
    saved = load_semiduplicate_approvals(mapping_path, table_section)

    current = {
        str(row["excel_columna"]): str(row["tabla_columna_propuesta"])
        for _, row in warnings_df.iterrows()
    }

    return current != saved


def load_mapping_store(mapping_path: Path) -> configparser.ConfigParser:
    cp = configparser.ConfigParser()
    cp.optionxform = str
    if mapping_path.exists():
        cp.read(mapping_path, encoding="utf-8")
    return cp


def get_stored_table_map(mapping_cp: configparser.ConfigParser, table_section: str) -> Dict[str, str]:
    out: Dict[str, str] = {}
    if table_section in mapping_cp:
        for src, target in mapping_cp[table_section].items():
            out[canonicalize_header(src)] = target.strip()
    return out


def find_best_target_column(raw_key: str, target_columns: List[str], threshold: float) -> tuple[str | None, float]:
    best_col = None
    best_score = 0.0
    for tgt in target_columns:
        score = difflib.SequenceMatcher(None, raw_key, canonicalize_header(tgt)).ratio()
        if score > best_score:
            best_score = score
            best_col = tgt
    if best_col and best_score >= threshold:
        return best_col, best_score
    return None, best_score


def normalize_excel_header_with_index(
    header: str,
    counter: Dict[str, int],
    raw_headers: List[str],
) -> str:
    txt = str(header).strip()

    # caso .1, .2, etc. -> siempre duplicado real de pandas
    dot_match = re.search(r"\.(\d+)$", txt)
    if dot_match:
        base = re.sub(r"\.\d+$", "", txt).strip()
        idx = int(dot_match.group(1)) + 1
        return f"{base}{idx}"

    # caso Estado4 / Detalle2 -> solo si existe también la base sin número
    num_match = re.search(r"^(.*?)(\d+)$", txt)
    if num_match:
        base = num_match.group(1).strip()
        idx = int(num_match.group(2))

        if idx >= 2:
            raw_set = {str(h).strip() for h in raw_headers}
            if base in raw_set:
                return f"{base}{idx}"

    # comportamiento normal por repeticiones exactas
    base_key = canonicalize_header(txt)
    counter.setdefault(base_key, 0)
    counter[base_key] += 1

    if counter[base_key] == 1:
        return txt
    return f"{txt}{counter[base_key]}"


def is_excel_duplicate_variant(header: str) -> bool:
    return bool(re.search(r"\.\d+$", str(header).strip()))


def get_semiduplicate_signature(header: str, raw_headers: List[str]) -> tuple[str, str] | None:
    txt = str(header).strip()

    # caso Estado.1 / Descripción.2
    m_dot = re.match(r"^(.*)\.(\d+)$", txt)
    if m_dot:
        base = m_dot.group(1).strip()
        idx = int(m_dot.group(2)) + 1
        return (canonicalize_header(base), f"{base}{idx}")

    # caso Estado4 / Detalle2 SOLO si existe la base
    m_num = re.match(r"^(.*?)(\d+)$", txt)
    if m_num:
        base = m_num.group(1).strip()
        idx = int(m_num.group(2))
        raw_set = {str(h).strip() for h in raw_headers}

        if idx >= 2 and base in raw_set:
            return (canonicalize_header(base), f"{base}{idx}")

    return None


def collect_semiduplicate_warnings(raw_headers: List[str], mapping_df: pd.DataFrame) -> pd.DataFrame:
    rows = []

    map_dict = dict(zip(mapping_df["excel_columna"], mapping_df["tabla_columna_propuesta"]))

    for raw in raw_headers:
        sig = get_semiduplicate_signature(raw, raw_headers)
        if not sig:
            continue

        base_key, normalized = sig
        proposed = map_dict.get(raw, "")

        rows.append(
            {
                "excel_columna": raw,
                "base_normalizada": base_key,
                "nombre_detectado": normalized,
                "tabla_columna_propuesta": proposed,
            }
        )

    if not rows:
        return pd.DataFrame(columns=[
            "excel_columna",
            "base_normalizada",
            "nombre_detectado",
            "tabla_columna_propuesta",
        ])

    return pd.DataFrame(rows)


def propose_header_mapping(
    raw_headers: List[str],
    target_columns: List[str],
    cfg_map: Dict[str, str],
    stored_map: Dict[str, str],
    similarity_threshold: float,
) -> pd.DataFrame:
    target_key_map = {canonicalize_header(col): col for col in target_columns}
    rows = []
    used_targets = set()
    header_counter = {}

    for raw_col in raw_headers:
        normalized_col = normalize_excel_header_with_index(raw_col, header_counter, raw_headers)
        
        raw_key = canonicalize_header(normalized_col)
        raw_base_key = raw_key

        mapped = None
        method = ""
        score = None

        # 1) prioridad absoluta: config.ini
        if raw_key in cfg_map:
            mapped = cfg_map[raw_key]
            method = "config_map"
            score = 1.0

        # 2) prioridad absoluta: mapping.ini
        elif raw_key in stored_map:
            mapped = stored_map[raw_key]
            method = "mapping_ini"
            score = 1.0

        # 3) match exacto por nombre completo
        elif raw_key in target_key_map and target_key_map[raw_key] not in used_targets:
            mapped = target_key_map[raw_key]
            method = "exact"
            score = 1.0

        # 4) si es columna duplicada tipo .1/.2 y la base ya fue usada, ignorar
        elif is_excel_duplicate_variant(raw_col):
            base_target = target_key_map.get(raw_base_key)
            if base_target and base_target in used_targets:
                mapped = IGNORE_COLUMN
                method = "excel_extra_duplicate"
                score = 0.0
            else:
                mapped, score = find_best_target_column(raw_key, target_columns, similarity_threshold)
                if mapped and mapped not in used_targets:
                    method = "fuzzy"
                else:
                    mapped = IGNORE_COLUMN
                    method = "excel_extra_duplicate"
                    score = 0.0

        # 5) fuzzy normal solo si no está ya usado
        else:
            candidate, candidate_score = find_best_target_column(raw_key, target_columns, similarity_threshold)
            if candidate and candidate not in used_targets:
                mapped = candidate
                method = "fuzzy"
                score = candidate_score
            else:
                mapped = IGNORE_COLUMN
                method = "excel_extra"
                score = 0.0

        if mapped != IGNORE_COLUMN:
            used_targets.add(mapped)

        rows.append(
            {
                "excel_columna": raw_col,
                "excel_normalizada": to_snake_name(raw_col),
                "tabla_columna_propuesta": mapped,
                "metodo": method,
                "score": round(float(score), 4) if score is not None else None,
            }
        )

    return pd.DataFrame(rows)


def apply_mapping_to_dataframe(df: pd.DataFrame, mapping_df: pd.DataFrame) -> pd.DataFrame:
    df = df.copy()
    col_map = dict(zip(mapping_df["excel_columna"], mapping_df["tabla_columna_propuesta"]))

    mapped_cols: List[str] = []
    used = set()
    cols_to_keep = []

    for raw_col in df.columns:
        mapped = col_map.get(raw_col, IGNORE_COLUMN)

        if mapped == IGNORE_COLUMN:
            continue

        if mapped in used:
            mapped = f"{mapped}__dup"

        used.add(mapped)
        mapped_cols.append(mapped)
        cols_to_keep.append(raw_col)

    df = df[cols_to_keep].copy()
    df.columns = mapped_cols
    return df


def save_mapping_ini(mapping_path: Path, table_section: str, mapping_df: pd.DataFrame) -> None:
    cp = load_mapping_store(mapping_path)
    if table_section not in cp:
        cp[table_section] = {}

    for _, row in mapping_df.iterrows():
        raw = str(row["excel_columna"])
        target = str(row["tabla_columna_propuesta"])
    
        if re.fullmatch(r"unnamed[:\s_0-9\-]*", raw.strip().lower()):
            continue
    
        if target == IGNORE_COLUMN:
            continue
    
        cp[table_section][raw] = target

    with mapping_path.open("w", encoding="utf-8") as f:
        cp.write(f)


def export_mapping_review(mapping_df: pd.DataFrame, output_dir: Path, schema: str, table: str) -> Path:
    output_dir.mkdir(parents=True, exist_ok=True)
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    out = output_dir / f"mapping_review_{schema}_{table}_{ts}.xlsx"

    with pd.ExcelWriter(out, engine="openpyxl") as writer:
        mapping_df.to_excel(writer, index=False)

    return out


def confirm_mapping(mapping_df: pd.DataFrame, mapping_path: Path, review_path: Path) -> str:
    low_conf = mapping_df[
        (mapping_df["score"] < 1) &
        (mapping_df["tabla_columna_propuesta"] != IGNORE_COLUMN)
    ]

    rel_mapping = mapping_path.relative_to(mapping_path.parent)
    rel_review = review_path.relative_to(mapping_path.parent)

    if low_conf.empty:
        print(">> [HOMOLOGACION] Todas las columnas correctamente homologadas ✔")
        print(f"\n>> [INFO] mapping.ini: \\{rel_mapping}")
        print(f">> [INFO] reporte excel: \\{rel_review}\n")
        return "yes"

    print("\n>> [HOMOLOGACION] Columnas a revisar (el resto está correctamente homologado):")
    print(low_conf.to_string(index=False))
    print(f"\n>> [INFO] mapping.ini: \\{rel_mapping}")
    print(f">> [INFO] reporte excel: \\{rel_review}\n")

    while True:
        ans = input("¿El mapeo es correcto? [s/n/r] (s=si, n=no, r=recargar): ").strip().lower()
        if ans in {"si", "s", "yes", "y"}:
            return "yes"
        if ans in {"no", "n"}:
            return "no"
        if ans in {"recargar", "r", "reload"}:
            return "reload"
        print("Respuesta no válida. Usa: s (si) / n (no) / r (recargar)")


def should_skip_mapping_confirmation(mapping_df: pd.DataFrame) -> bool:
    if mapping_df.empty:
        return False

    effective_df = mapping_df[mapping_df["tabla_columna_propuesta"] != IGNORE_COLUMN].copy()
    if effective_df.empty:
        return False

    allowed = {"mapping_ini", "config_map", "exact"}
    methods_ok = effective_df["metodo"].isin(allowed).all()
    no_dup_ok = ~effective_df["excel_normalizada"].astype("string").str.contains("__dup", na=False).any()
    no_error_col_ok = ~effective_df["tabla_columna_propuesta"].astype("string").eq("errores").any()

    return bool(methods_ok and no_dup_ok and no_error_col_ok)


def confirm_semiduplicate_warnings(
    warnings_df: pd.DataFrame,
    mapping_path: Path,
    table_section: str,
    non_interactive: bool,
) -> None:
    if warnings_df.empty:
        return

    changed = semiduplicate_warnings_changed(mapping_path, table_section, warnings_df)

    print("\n>> [WARNING] Columnas semi-duplicadas detectadas:")
    print(warnings_df[["excel_columna", "tabla_columna_propuesta"]].to_string(index=False))

    if not changed:
        print(">> [WARNING] Ya estaban aprobadas previamente ✔")
        return

    if non_interactive:
        print(">> [WARNING] Aprobadas automáticamente por --non-interactive")
        save_semiduplicate_approvals(mapping_path, table_section, warnings_df)
        return

    while True:
        ans = input("¿Apruebas estas homologaciones semi-duplicadas? [s/n]: ").strip().lower()
        if ans in {"s", "si", "y", "yes"}:
            save_semiduplicate_approvals(mapping_path, table_section, warnings_df)
            print("\n>> [WARNING] Semi-duplicadas aprobadas y guardadas ✔")
            return
        if ans in {"n", "no"}:
            raise UserCancelledLoad(
                "Carga detenida por usuario: las columnas semi-duplicadas requieren revisión."
            )
        print("Respuesta no válida. Usa: s (si) / n (no)")


def resolve_mapping(
    df_raw: pd.DataFrame,
    cfg: configparser.ConfigParser,
    mapping_path: Path,
    output_dir: Path,
    schema: str,
    table: str,
    target_columns: List[str],
    similarity_threshold: float,
    auto_approve: bool,
    non_interactive: bool,
    auto_confirm_known_mapping: bool,
    cleanup_mapping_review: bool,
) -> tuple[pd.DataFrame, Path | None, pd.DataFrame]:
    cfg_map = get_config_column_map(cfg)
    table_section = f"{schema}.{table}"

    stored_map = get_stored_table_map(load_mapping_store(mapping_path), table_section)
    
    mapping_df = propose_header_mapping(
        raw_headers=[str(c) for c in df_raw.columns],
        target_columns=target_columns,
        cfg_map=cfg_map,
        stored_map=stored_map,
        similarity_threshold=similarity_threshold,
    )

    semidup_warnings_df = collect_semiduplicate_warnings(
        raw_headers=[str(c) for c in df_raw.columns],
        mapping_df=mapping_df,
    )

    review_path = export_mapping_review(mapping_df, output_dir, schema, table)
    save_mapping_ini(mapping_path, table_section, mapping_df)

    skip_confirm = auto_confirm_known_mapping and should_skip_mapping_confirmation(mapping_df)

    if auto_approve or non_interactive or skip_confirm:
        decision = "yes"
        if auto_approve:
            print("\n>> [HOMOLOGACION] Auto-aprobada por --auto-approve-mapping")
        elif non_interactive:
            print("\n>> [HOMOLOGACION] Auto-aprobada por --non-interactive")
        else:
            if (mapping_df["metodo"] == "mapping_ini").all():
                print("\n>> [HOMOLOGACION] Mapeo reutilizado desde mapping.ini")
            else:
                print("\n>> [HOMOLOGACION] Mapeo generado automáticamente y validado")
    else:
        decision = confirm_mapping(mapping_df, mapping_path, review_path)
        while decision == "reload":
            print("\n\n>> [HOMOLOGACION] Recargando mapping.ini actualizado...")
            stored_map = get_stored_table_map(load_mapping_store(mapping_path), table_section)
            mapping_df = propose_header_mapping(
                raw_headers=[str(c) for c in df_raw.columns],
                target_columns=target_columns,
                cfg_map=cfg_map,
                stored_map=stored_map,
                similarity_threshold=similarity_threshold,
            )
            semidup_warnings_df = collect_semiduplicate_warnings(
                raw_headers=[str(c) for c in df_raw.columns],
                mapping_df=mapping_df,
            )
            review_path = export_mapping_review(mapping_df, output_dir, schema, table)
            save_mapping_ini(mapping_path, table_section, mapping_df)
            decision = confirm_mapping(mapping_df, mapping_path, review_path)

    if decision == "no":
        raise UserCancelledLoad("Carga detenida por usuario. Ajusta mapping.ini o la estructura del Excel y vuelve a ejecutar cuando quieras.")

    confirm_semiduplicate_warnings(
        warnings_df=semidup_warnings_df,
        mapping_path=mapping_path,
        table_section=table_section,
        non_interactive=non_interactive,
    )

    if cleanup_mapping_review:
        safe_delete(review_path)
        review_path = None

    df_mapped = apply_mapping_to_dataframe(df_raw, mapping_df)
    return df_mapped, review_path, mapping_df


# =========================================================
# PARSERS
# =========================================================

def parse_excel_serial_dates(series: pd.Series, existing: pd.Series | None = None) -> pd.Series:
    s = series.astype("string").str.strip()
    parsed = existing.copy() if existing is not None else pd.Series(pd.NaT, index=series.index, dtype="datetime64[ns]")

    missing = parsed.isna() & s.notna() & (s != "")
    if not missing.any():
        return parsed

    numeric_candidate = pd.to_numeric(s[missing], errors="coerce")
    serial_mask = numeric_candidate.notna() & (numeric_candidate >= 1) & (numeric_candidate <= 73050)
    if serial_mask.any():
        serial_vals = numeric_candidate[serial_mask]
        parsed_serial = pd.to_datetime(serial_vals, unit="D", origin="1899-12-30", errors="coerce")
        parsed.loc[parsed_serial.index] = parsed_serial

    return parsed


def parse_periodo_series(series: pd.Series) -> pd.Series:
    s = series.astype("string").str.strip().str.lower()
    parsed = pd.to_datetime(s, errors="coerce", dayfirst=True)

    missing = parsed.isna() & s.notna() & (s != "")
    if missing.any():
        sub = s[missing]
        pattern = re.compile(r"^(ene|feb|mar|abr|may|jun|jul|ago|sep|oct|nov|dic)[\-\/]?(\d{2}|\d{4})$")

        built = []
        idxs = []
        for idx, val in sub.items():
            m = pattern.fullmatch(str(val))
            if not m:
                continue
            mon_txt, year_txt = m.group(1), m.group(2)
            month = SPANISH_MONTHS[mon_txt]
            year = int(year_txt)
            if year < 100:
                year += 2000
            try:
                built.append(pd.Timestamp(year=year, month=month, day=1))
                idxs.append(idx)
            except Exception:
                pass

        if idxs:
            parsed.loc[idxs] = pd.Series(built, index=idxs)

    return parsed


def parse_numeric_series(series: pd.Series) -> pd.Series:
    s = series.astype("string").str.strip()

    both = s.str.contains(r"\.", regex=True, na=False) & s.str.contains(",", regex=True, na=False)
    comma_only = ~both & s.str.contains(",", regex=True, na=False)

    s = s.where(~both, s.str.replace(".", "", regex=False).str.replace(",", ".", regex=False))
    s = s.where(~comma_only, s.str.replace(",", ".", regex=False))

    s = s.str.replace(r"^np\.float64\((.+)\)$", r"\1", regex=True)
    s = s.str.replace(r"^np\.int64\((.+)\)$", r"\1", regex=True)

    return pd.to_numeric(s, errors="coerce")


def parse_bool_series(series: pd.Series) -> pd.Series:
    return (
        series.astype("string")
        .str.strip()
        .str.lower()
        .replace(
            {
                "true": True,
                "false": False,
                "1": True,
                "0": False,
                "si": True,
                "sí": True,
                "no": False,
                "<na>": pd.NA,
                "nan": pd.NA,
                "": pd.NA,
            }
        )
    )


# =========================================================
# VALIDATION
# =========================================================

def validate_and_transform(
    df_raw: pd.DataFrame,
    metadata: List[ColumnMeta],
    insert_cols: List[str],
    fixed_cols: List[str],
    min_year: int,
    max_year: int,
    allow_out_of_range_dates: bool = False,
) -> ValidationResult:
    df = df_raw.copy()
    input_cols = [c for c in insert_cols if c not in fixed_cols]

    for col in input_cols:
        if col not in df.columns:
            df[col] = pd.NA

    df = df[input_cols].copy()

    for col in fixed_cols:
        if col not in df.columns:
            df[col] = pd.NA

    type_map = {c.name: c.data_type for c in metadata}
    errors: List[pd.Series] = []

    for col in input_cols:
        data_type = type_map.get(col, "text")
        original = df[col]
           
        if data_type in DATE_TYPES or data_type in TIMESTAMP_TYPES:
            if allow_out_of_range_dates and data_type in DATE_TYPES:
                parsed = parse_date_series_unbounded(original)
                bad = original.notna() & (original.astype(str).str.strip() != "") & parsed.isna()
                errors.append(bad.rename(f"error_{col}"))
                df[col] = parsed.where(~bad, pd.NA)
            else:
                parsed = parse_periodo_series(original) if col == "periodo" else pd.to_datetime(original, errors="coerce", dayfirst=True)
                parsed = parse_excel_serial_dates(original, parsed)
        
                bad_parse = original.notna() & (original.astype(str).str.strip() != "") & parsed.isna()
                bad_year = parsed.notna() & ((parsed.dt.year < min_year) | (parsed.dt.year > max_year))
                bad = bad_parse | bad_year
        
                errors.append(bad.rename(f"error_{col}"))
        
                if data_type in DATE_TYPES:
                    df[col] = parsed.where(~bad, pd.NaT).dt.date
                else:
                    df[col] = parsed.where(~bad, pd.NaT)

        elif data_type in NUMERIC_TYPES:
            parsed = parse_numeric_series(original)
            bad = original.notna() & (original.astype(str).str.strip() != "") & parsed.isna()
            errors.append(bad.rename(f"error_{col}"))
            df[col] = parsed

        elif data_type in BOOL_TYPES:
            normalized = parse_bool_series(original)
            bad = normalized.notna() & ~normalized.isin([True, False])
            errors.append(bad.rename(f"error_{col}"))
            df[col] = normalized.where(normalized.isin([True, False]), pd.NA)

    if "fecha_inicio_proyecto" in df.columns and "fecha_fin_proyecto" in df.columns:
        start = pd.to_datetime(df["fecha_inicio_proyecto"], errors="coerce")
        end = pd.to_datetime(df["fecha_fin_proyecto"], errors="coerce")
        bad_date_order = start.notna() & end.notna() & (start > end)
        errors.append(bad_date_order.rename("error_rango_fechas_proyecto"))

    error_df = pd.concat(errors, axis=1) if errors else pd.DataFrame(index=df.index)
    row_has_error = error_df.any(axis=1) if not error_df.empty else pd.Series(False, index=df.index)

    error_messages = pd.Series("", index=df.index, dtype="string")
    invalid_df = df_raw.loc[row_has_error].copy()

    if not error_df.empty:
        invalid_msgs = error_df.loc[row_has_error].apply(
            lambda row: ", ".join([c.replace("error_", "") for c, v in row.items() if bool(v)]),
            axis=1,
        )
        invalid_df["errores"] = invalid_msgs
        error_messages.loc[invalid_msgs.index] = invalid_msgs.astype("string")

    valid_df = df.loc[~row_has_error].copy()
    return ValidationResult(valid_df=valid_df, invalid_df=invalid_df, error_messages=error_messages)


# =========================================================
# DEFAULTS / FIXED VALUES
# =========================================================

def parse_column_default_literal(default_expr: str | None) -> tuple[bool, Any]:
    if default_expr is None:
        return False, None

    expr = str(default_expr).strip()
    if not expr:
        return False, None

    expr_no_cast = re.sub(r"::[a-zA-Z0-9_\s\[\]\.\"]+$", "", expr).strip()
    low = expr_no_cast.lower()

    if low in {"true", "false"}:
        return True, low == "true"

    if re.fullmatch(r"[-+]?\d+", expr_no_cast):
        return True, int(expr_no_cast)

    if re.fullmatch(r"[-+]?\d*\.\d+", expr_no_cast):
        return True, float(expr_no_cast)

    m_text = re.fullmatch(r"'((?:''|[^'])*)'", expr_no_cast)
    if m_text:
        return True, m_text.group(1).replace("''", "'")

    return False, None


def build_missing_column_suggestions_from_ignored(
    ignored_df: pd.DataFrame,
    missing_plan: List[Dict[str, Any]],
    similarity_threshold: float,
) -> pd.DataFrame:
    rows = []

    ignored_excel_cols = ignored_df["excel_columna"].astype(str).tolist()
    one_to_one_hint = len(ignored_excel_cols) == 1 and len(missing_plan) == 1

    for item in missing_plan:
        target_col = item["column"]

        best_excel = None
        best_score = 0.0

        for excel_col in ignored_excel_cols:
            excel_key = canonicalize_header(excel_col)
            target_key = canonicalize_header(target_col)

            score = difflib.SequenceMatcher(None, excel_key, target_key).ratio()

            if excel_key and target_key:
                if excel_key in target_key or target_key in excel_key:
                    score = max(score, 0.75)

            if score > best_score:
                best_score = score
                best_excel = excel_col

        sugerencia_util = (best_score >= similarity_threshold) or one_to_one_hint

        rows.append(
            {
                "tabla_columna_faltante": target_col,
                "excel_columna_ignorada": best_excel if best_score > 0 or one_to_one_hint else "",
                "score_sugerido": round(best_score, 4),
                "sugerencia_util": sugerencia_util,
            }
        )

    return pd.DataFrame(rows)


def print_mapping_coverage_summary(
    df_raw: pd.DataFrame,
    mapping_df: pd.DataFrame,
    insert_cols: List[str],
    fixed_cols: List[str],
    missing_plan: List[Dict[str, Any]],
    similarity_threshold: float,
) -> pd.DataFrame:
    total_excel_cols = len(df_raw.columns)

    mapped_effective = mapping_df[
        mapping_df["tabla_columna_propuesta"].astype("string").notna() &
        (mapping_df["tabla_columna_propuesta"].astype("string") != IGNORE_COLUMN)
    ]

    ignored_cols = mapping_df[
        mapping_df["tabla_columna_propuesta"].astype("string") == IGNORE_COLUMN
    ].copy()

    print("\n>> [VALIDACION] Cobertura de columnas:")
    print(f"   - columnas excel leídas: {total_excel_cols}")
    print(f"   - columnas homologadas a tabla: {len(mapped_effective)}")
    print(f"   - columnas extra ignoradas: {len(ignored_cols)}")
    print(f"   - columnas de tabla faltantes: {len(missing_plan)}")

    if not ignored_cols.empty:
        print("\n>> [VALIDACION] Columnas ignoradas del Excel:")
        print(ignored_cols[["excel_columna"]].to_string(index=False))

    suggestions_df = pd.DataFrame()

    if missing_plan:
        print("\n>> [VALIDACION] Columnas de tabla faltantes en Excel:")
        for item in missing_plan:
            col = item["column"]
            default_expr = item["default_expr"]
            has_default = item["has_literal_default"]

            if has_default:
                msg = f"DEFAULT literal ({default_expr})"
            elif default_expr:
                msg = f"DEFAULT no literal ({default_expr})"
            else:
                msg = "sin DEFAULT"

            print(f"   - {col}: {msg}")

        suggestions_df = build_missing_column_suggestions_from_ignored(
            ignored_df=ignored_cols,
            missing_plan=missing_plan,
            similarity_threshold=similarity_threshold,
        )

        suggestions_df = suggestions_df[
            suggestions_df["excel_columna_ignorada"].astype("string").str.strip() != ""
        ]

        if not suggestions_df.empty:
            print("\n>> [VALIDACION] Posibles homologaciones sugeridas (revisar mapping.ini si aplica):")
            print(
                suggestions_df[
                    ["excel_columna_ignorada", "tabla_columna_faltante", "score_sugerido", "sugerencia_util"]
                ].to_string(index=False)
            )

    return suggestions_df


def collect_missing_input_columns(
    df: pd.DataFrame,
    metadata: List[ColumnMeta],
    insert_cols: List[str],
    fixed_cols: List[str],
) -> List[Dict[str, Any]]:
    present = set(df.columns)
    meta_map = {m.name: m for m in metadata}
    missing = [c for c in insert_cols if c not in fixed_cols and c not in present]

    out: List[Dict[str, Any]] = []
    for col in missing:
        meta = meta_map.get(col)
        default_expr = meta.column_default if meta else None
        has_literal_default, default_value = parse_column_default_literal(default_expr)
        out.append(
            {
                "column": col,
                "default_expr": default_expr,
                "has_literal_default": has_literal_default,
                "default_value": default_value,
            }
        )
    return out


def confirm_missing_columns_plan(
    missing_plan: List[Dict[str, Any]],
    suggestions_df: pd.DataFrame,
    mapping_path: Path,
    table_section: str,
    non_interactive: bool,
) -> str:
    if not missing_plan:
        return "continue"

    useful_suggestions = suggestions_df[suggestions_df["sugerencia_util"]].copy()

    if not useful_suggestions.empty and not non_interactive:
        while True:
            ans = input(
                "¿Agregar las sugerencias útiles a mapping.ini y recargar? [s/n] "
                "(s=guardar y recargar, n=seguir evaluando): "
            ).strip().lower()

            if ans in {"s", "si", "yes", "y"}:
                save_missing_column_suggestions_to_mapping(
                    mapping_path=mapping_path,
                    table_section=table_section,
                    suggestions_df=useful_suggestions,
                )
                return "reload"

            if ans in {"n", "no"}:
                break

            print("Respuesta no válida. Usa: s (guardar y recargar) / n (seguir evaluando)")

    while True:
        ans = input(
            "¿Continuar con columnas faltantes usando DEFAULT/NULL? [s/n] "
            "(s=aceptar, n=cancelar carga): "
        ).strip().lower()

        if ans in {"s", "si", "yes", "y"}:
            return "continue"

        if ans in {"n", "no"}:
            return "cancel"

        print("Respuesta no válida. Usa: s (aceptar) / n (cancelar carga)")


def apply_missing_columns_plan(df: pd.DataFrame, missing_plan: List[Dict[str, Any]]) -> pd.DataFrame:
    if not missing_plan:
        return df

    out = df.copy()
    for item in missing_plan:
        col = item["column"]
        out[col] = item["default_value"] if item["has_literal_default"] else pd.NA
    return out


def parse_fixed_value(token: str) -> Any:
    val = str(token).strip()
    low = val.lower()

    if low == "__now_ts__":
        return datetime.now()
    if low == "__today__":
        return datetime.now().date()
    if low == "__true__":
        return True
    if low == "__false__":
        return False
    if low == "__null__":
        return None

    return val


def apply_fixed_values(df: pd.DataFrame, cfg: configparser.ConfigParser, insert_cols: List[str]) -> pd.DataFrame:
    df = df.copy()
    if "fixed_values" not in cfg:
        return df

    for col, token in cfg["fixed_values"].items():
        if col in insert_cols:
            df[col] = parse_fixed_value(token)
    return df


# =========================================================
# DB LOAD
# =========================================================

def execute_pre_import_action(conn_params: Dict[str, str], action_sql: str) -> bool:
    if not action_sql or not action_sql.strip():
        return True

    statement = action_sql.strip()
    try:
        with psycopg2.connect(**conn_params) as conn:
            with conn.cursor() as cur:
                cur.execute(statement)
            conn.commit()
        print(f"\n>> [IMPORT_ACTION] Ejecutada previa: {statement}")
        return True
    except Exception as e:
        print(f"\n>> [IMPORT_ACTION] Error ejecutando {statement}: {e}")
        return False


def to_db_value(value):
    if pd.isna(value):
        return None

    if hasattr(value, "item") and callable(getattr(value, "item")):
        try:
            value = value.item()
        except Exception:
            pass

    if isinstance(value, str):
        txt = value.strip()
        m_float = re.fullmatch(r"np\.float64\(([-+]?\d+(?:\.\d+)?(?:[eE][-+]?\d+)?)\)", txt)
        m_int = re.fullmatch(r"np\.int64\(([-+]?\d+)\)", txt)
        if m_float:
            return float(m_float.group(1))
        if m_int:
            return int(m_int.group(1))
        return txt

    if isinstance(value, float) and (math.isnan(value) or math.isinf(value)):
        return None

    if isinstance(value, pd.Timestamp):
        return value.to_pydatetime()

    return value


def insert_valid_rows(
    df: pd.DataFrame,
    conn_params: Dict[str, str],
    schema: str,
    table: str,
    insert_cols: List[str],
    batch_size: int,
    progress_every: int,
) -> int:
    if df.empty:
        return 0

    rows = [tuple(to_db_value(v) for v in row) for row in df[insert_cols].itertuples(index=False, name=None)]
    total = len(rows)

    insert_stmt = sql.SQL("INSERT INTO {}.{} ({}) VALUES %s").format(
        sql.Identifier(schema),
        sql.Identifier(table),
        sql.SQL(", ").join(sql.Identifier(c) for c in insert_cols),
    )

    inserted = 0
    next_progress_mark = progress_every if progress_every > 0 else None

    with psycopg2.connect(**conn_params) as conn:
        with conn.cursor() as cur:
            for start in range(0, total, batch_size):
                end = min(start + batch_size, total)
                execute_values(cur, insert_stmt.as_string(conn), rows[start:end], page_size=batch_size)
                inserted = end

                while next_progress_mark is not None and inserted >= next_progress_mark:
                    print(f">> [PROGRESO] Insertados {next_progress_mark} de {total} registros")
                    next_progress_mark += progress_every

        conn.commit()

    if progress_every > 0 and inserted % progress_every != 0:
        print(f"\n>> [PROGRESO] Insertados {inserted} de {total} registros")

    return inserted


# =========================================================
# RETRY / OUTPUT
# =========================================================

def export_invalid(invalid_df: pd.DataFrame, output_dir: Path) -> Path | None:
    if invalid_df.empty:
        return None
    output_dir.mkdir(parents=True, exist_ok=True)
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    out = output_dir / f"registros_invalidos_{ts}.xlsx"

    with pd.ExcelWriter(out, engine="openpyxl") as writer:
        invalid_df.to_excel(writer, index=False)

    return out


def copy_invalid_to_retry(invalid_path: Path | None, retry_input_dir: Path) -> Path | None:
    if not invalid_path:
        return None
    retry_input_dir.mkdir(parents=True, exist_ok=True)
    destination = retry_input_dir / invalid_path.name
    if destination.exists():
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        destination = retry_input_dir / f"{invalid_path.stem}_{ts}{invalid_path.suffix}"
    shutil.copy2(invalid_path, destination)
    return destination


def load_retry_index(index_path: Path) -> Dict[str, Dict[str, str]]:
    if not index_path.exists():
        return {}
    try:
        with index_path.open("r", encoding="utf-8") as f:
            data = json.load(f)
            return data if isinstance(data, dict) else {}
    except Exception:
        return {}


def save_retry_index(index_path: Path, data: Dict[str, Dict[str, str]]) -> None:
    index_path.parent.mkdir(parents=True, exist_ok=True)
    with index_path.open("w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)


def register_retry_entry(
    index_path: Path,
    retry_file: Path,
    original_processed_path: Path | None,
    invalid_report_path: Path | None,
) -> None:
    db = load_retry_index(index_path)
    db[retry_file.name] = {
        "original_processed_path": str(original_processed_path) if original_processed_path else "",
        "invalid_report_path": str(invalid_report_path) if invalid_report_path else "",
    }
    save_retry_index(index_path, db)


def pop_retry_entry(index_path: Path, retry_file: Path) -> Dict[str, str] | None:
    db = load_retry_index(index_path)
    entry = db.pop(retry_file.name, None)
    save_retry_index(index_path, db)
    return entry


def annotate_source_excel_errors(
    excel_path: Path,
    sheet_name: str,
    error_messages: pd.Series,
    header_row_idx: int,
) -> bool:
    if error_messages.empty:
        return False

    if excel_path.suffix.lower() != ".xlsx":
        return False

    try:
        wb = load_workbook(excel_path)
        if sheet_name not in wb.sheetnames:
            wb.close()
            print(f"\n>> [INFO] Hoja '{sheet_name}' no encontrada para anotar errores en {excel_path.name}.")
            return False

        ws = wb[sheet_name]
        excel_header_row = header_row_idx + 1
        last_col = ws.max_column
        error_col = None

        for col in range(1, last_col + 1):
            val = ws.cell(row=excel_header_row, column=col).value
            if str(val).strip().lower() == "errores":
                error_col = col
                break

        if error_col is None:
            error_col = last_col + 1
            ws.cell(row=excel_header_row, column=error_col, value="errores")

        max_row = ws.max_row
        first_data_row = excel_header_row + 1

        for r in range(first_data_row, max_row + 1):
            ws.cell(row=r, column=error_col, value=None)

        for idx, msg in error_messages.dropna().items():
            txt = str(msg).strip()
            if not txt:
                continue
            excel_row = int(idx) + header_row_idx + 2
            if excel_row <= max_row:
                ws.cell(row=excel_row, column=error_col, value=txt)

        wb.save(excel_path)
        wb.close()
        return True
    except Exception as e:
        print(f"[INFO] No se pudo anotar errores en excel fuente: {e}")
        return False


def export_annotated_xlsx_from_source(
    df_raw: pd.DataFrame,
    error_messages: pd.Series,
    output_path: Path,
) -> Path:
    out_df = df_raw.copy()

    if "errores" in out_df.columns:
        out_df = out_df.drop(columns=["errores"])

    out_df["errores"] = error_messages.reindex(out_df.index).fillna("").astype("string")
    output_path.parent.mkdir(parents=True, exist_ok=True)

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        out_df.to_excel(writer, index=False)

    return output_path

def mark_excel_as_processed(
    excel_path: Path,
    mode: str,
    done_dir: Path,
    loaded_suffix: str = "_LOADED",
    status_suffix: str | None = None,
    custom_name_suffix: str | None = None,
) -> Path | None:
    mode = (mode or "none").strip().lower()
    if mode == "none":
        return None

    ts = datetime.now().strftime("%Y%m%d_%H%M%S")

    if mode == "move":
        done_dir.mkdir(parents=True, exist_ok=True)

        if custom_name_suffix:
            destination_name = f"{excel_path.stem}{custom_name_suffix}{excel_path.suffix}"
        else:
            suffix = f"{loaded_suffix}{status_suffix or ''}"
            destination_name = f"{excel_path.stem}{suffix}{excel_path.suffix}"

        destination = done_dir / destination_name

        if destination.exists():
            destination = done_dir / f"{Path(destination_name).stem}_{ts}{excel_path.suffix}"

        excel_path.replace(destination)
        return destination

    if mode == "rename":
        if custom_name_suffix:
            destination = excel_path.with_name(f"{excel_path.stem}{custom_name_suffix}{excel_path.suffix}")
        else:
            suffix = f"{loaded_suffix}{status_suffix or ''}"
            destination = excel_path.with_name(f"{excel_path.stem}{suffix}{excel_path.suffix}")

        if destination.exists():
            fallback_suffix = custom_name_suffix or loaded_suffix
            destination = excel_path.with_name(f"{excel_path.stem}{fallback_suffix}_{ts}{excel_path.suffix}")

        excel_path.replace(destination)
        return destination

    raise ValueError("processed_mode debe ser: none, move o rename")


def rename_partial_to_ok(original_processed_path: Path) -> Path | None:
    if not original_processed_path.exists():
        return None

    stem = original_processed_path.stem
    if "_PARTIAL_ERROR" in stem:
        new_stem = stem.replace("_PARTIAL_ERROR", "_OK")
    elif stem.endswith("_OK"):
        return original_processed_path
    else:
        new_stem = f"{stem}_OK"

    candidate = original_processed_path.with_name(f"{new_stem}{original_processed_path.suffix}")
    if candidate.exists():
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        candidate = original_processed_path.with_name(f"{new_stem}_{ts}{original_processed_path.suffix}")
    original_processed_path.replace(candidate)
    return candidate


# =========================================================
# CLI
# =========================================================

def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description="Carga masiva genérica Excel -> PostgreSQL")
    parser.add_argument("--config-path", help="Ruta a config.ini (por defecto, junto al script)")
    parser.add_argument("--auto-approve-mapping", action="store_true", help="Aprueba mapeo automáticamente sin interacción")
    parser.add_argument("--only-mapping", action="store_true", help="Genera/valida homologación y sale sin insertar datos")
    parser.add_argument("--interactive-target", action="store_true", help="Permite seleccionar schema/tabla en consola")
    parser.add_argument(
        "--target-section",
        choices=["target", "target_defensa"],
        help="target usa [target]; target_defensa usa [target_defensa_options]",
    )
    parser.add_argument("--load-mode", choices=["initial", "retry"], help="Modo de carga sin prompt")
    parser.add_argument("--yes-missing-columns", action="store_true", help="Acepta columnas faltantes sin prompt")
    parser.add_argument("--non-interactive", action="store_true", help="Modo no interactivo")
    parser.add_argument("--log-file", help="Nombre/ruta de log. Si se omite, se genera en ./logs automáticamente")
    parser.add_argument("--run-tests", action="store_true", help="Ejecuta pruebas mínimas y sale")
    return parser


def resolve_log_file(args, script_dir: Path) -> Path:
    logs_dir = (script_dir / "logs").resolve()
    ts_run = datetime.now().strftime("%Y%m%d_%H%M%S")

    if args.log_file:
        raw_log = Path(args.log_file)
        if raw_log.is_absolute():
            return raw_log
        if raw_log.parent == Path("."):
            return (logs_dir / raw_log.name).resolve()
        return (script_dir / raw_log).resolve()

    return logs_dir / f"run_{ts_run}.log"


def choose_target_defensa(cfg: configparser.ConfigParser) -> tuple[str, str]:
    section = "target_defensa_options"
    if section not in cfg or not cfg[section]:
        raise KeyError("Falta sección [target_defensa_options] en config.ini o está vacía.")

    options_map = dict(cfg[section])

    print("\n[SELECCION] Tablas disponibles para Defensa")
    ordered_keys = sorted(options_map.keys(), key=lambda x: int(x) if str(x).isdigit() else str(x))

    for key in ordered_keys:
        full_name = options_map[key].strip()
        print(f"  {key}. {full_name}")

    while True:
        choice = input(f"Selecciona una opción ({'/'.join(ordered_keys)}): ").strip()
        if choice in options_map:
            full_name = options_map[choice].strip()
            if "." not in full_name:
                raise ValueError(
                    f"La opción {choice} en [target_defensa_options] debe tener formato schema.tabla"
                )
            schema, table = full_name.split(".", 1)
            return schema.strip(), table.strip()

        print("Selección inválida, intenta de nuevo.")


def confirm_or_update_default_target(
    cfg: configparser.ConfigParser,
    config_path: Path,
    conn_params: Dict[str, str],
) -> tuple[str, str]:
    """Confirma target actual de [target] o permite reemplazarlo por consola, validando schema y tabla."""
    if "target" not in cfg:
        raise KeyError("Falta sección [target] en config.ini")

    current_schema = (cfg["target"].get("schema") or "").strip()
    current_table = (cfg["target"].get("table") or "").strip()

    if current_schema and current_table:
        print(f"\n[SELECCION] Target actual configurado: {current_schema}.{current_table}")
        if ask_yes_no("¿El target es correcto?"):
            if not schema_exists(conn_params, current_schema):
                print(f"[VALIDACION] El schema '{current_schema}' no existe en la base de datos.")
            elif not table_exists(conn_params, current_schema, current_table):
                print(f"[VALIDACION] La tabla '{current_schema}.{current_table}' no existe en la base de datos.")
            else:
                return current_schema, current_table
    else:
        print("[INFO] La sección [target] no está completa en config.ini.")

    while True:
        raw = input("Ingresa el nuevo target en formato schema.tabla: ").strip()

        if raw.count(".") != 1:
            print("Formato inválido. Debe ser exactamente: schema.tabla")
            continue

        schema, table = raw.split(".", 1)
        schema = schema.strip()
        table = table.strip()

        if not schema or not table:
            print("Schema y tabla no pueden estar vacíos.")
            continue

        if not schema_exists(conn_params, schema):
            print(f"El schema '{schema}' no existe en la base de datos.")
            continue

        if not table_exists(conn_params, schema, table):
            print(f"La tabla '{schema}.{table}' no existe en la base de datos.")
            continue

        cfg["target"]["schema"] = schema
        cfg["target"]["table"] = table

        with config_path.open("w", encoding="utf-8") as f:
            cfg.write(f)

        print(f"[CONFIG] Target actualizado y guardado en config.ini: {schema}.{table}")
        return schema, table


# =========================================================
# MAIN
# =========================================================

def main() -> None:
    parser = build_parser()
    args = parser.parse_args()

    if args.run_tests:
        run_tests()
        return

    script_dir = Path(__file__).resolve().parent
    config_path = Path(args.config_path) if args.config_path else (script_dir / "config.ini")
    log_file = resolve_log_file(args, script_dir)

    setup_logging(log_file)
    try:
        relative_log = log_file.relative_to(script_dir)
        log_msg = f".\\{relative_log}"
    except ValueError:
        log_msg = str(log_file)
    
    print(f">> [LOG] Guardando ejecución en: {log_msg}\n")

    cfg = load_config(config_path)
    conn_params = get_db_params(cfg)

    if args.non_interactive and args.interactive_target:
        raise ValueError("--non-interactive no se puede combinar con --interactive-target")

    if args.interactive_target:
        schemas = fetch_available_schemas(conn_params)
        schema = prompt_choice("Schema destino", schemas)
        tables = fetch_tables_in_schema(conn_params, schema)
        table = prompt_choice(f"Tabla destino en schema '{schema}'", tables)
        print(f"\n>> [SELECCIONADO] Usando destino: {schema}.{table}")
    else:
        if args.target_section:
            if args.target_section == "target":
                if args.non_interactive:
                    schema = cfg["target"].get("schema")
                    table = cfg["target"].get("table")
                else:
                    schema, table = confirm_or_update_default_target(cfg, config_path, conn_params)
                print(f"\n>> [SELECCIONADO] Usando destino desde [target]: {schema}.{table}")

            elif args.target_section == "target_defensa":
                schema, table = choose_target_defensa(cfg)
                print(f"\n>> [SELECCIONADO] Usando destino Defensa: {schema}.{table}")
            else:
                raise ValueError(f"Sección de destino no soportada: {args.target_section}")
        else:
            if "target_defensa_options" in cfg and not args.non_interactive:
                is_defensa = ask_yes_no("¿El import masivo es para Defensa?")
                if is_defensa:
                    schema, table = choose_target_defensa(cfg)
                    print(f"\n>> [SELECCIONADO] Usando destino Defensa: {schema}.{table}")
                else:
                    schema, table = confirm_or_update_default_target(cfg, config_path, conn_params)
                    print(f"\n>> [SELECCIONADO] Usando destino desde [target]: {schema}.{table}")
            else:
                if args.non_interactive:
                    schema = cfg["target"].get("schema")
                    table = cfg["target"].get("table")
                else:
                    schema, table = confirm_or_update_default_target(cfg, config_path, conn_params)
                print(f"\n>> [SELECCIONADO] Usando destino desde [target]: {schema}.{table}")

    table_identifier = normalize_table_identifier(schema, table)

    ALLOW_OUT_OF_RANGE_DATE_TABLES = {
        "proyecto_dashboard_defensa.converge_proyectos_financieros",
    }
    
    allow_out_of_range_dates = table_identifier in ALLOW_OUT_OF_RANGE_DATE_TABLES

    input_dir = resolve_path(script_dir, cfg["input"].get("input_dir", "./inputs"))
    retry_input_dir = resolve_path(script_dir, cfg["input"].get("retry_input_dir", "./inputs_retry"))
    file_name = cfg["input"].get("file_name", fallback="").strip() or None
    sheet_name = cfg["input"].get("sheet_name", fallback="bbdd")

    output_dir = resolve_path(script_dir, cfg["output"].get("output_dir", "./salidas"))
    mapping_path = resolve_path(script_dir, cfg["output"].get("mapping_file", "mapping.ini"))
    processed_dir = resolve_path(script_dir, cfg["output"].get("processed_dir", "./excels_done"))
    retry_index_path = resolve_path(script_dir, cfg["output"].get("retry_index_file", "retry_index.json"))

    processed_mode = cfg["output"].get("processed_mode", fallback="none")
    loaded_suffix = cfg["output"].get("loaded_suffix", fallback="_LOADED")
    cleanup_mapping_review = cfg["output"].getboolean("cleanup_mapping_review", fallback=True)

    auto_confirm_known_mapping = cfg["run"].getboolean("auto_confirm_known_mapping", fallback=True)
    batch_size = cfg["run"].getint("batch_size", fallback=1000)
    progress_every = cfg["run"].getint("progress_every", fallback=10000)
    min_year = cfg["run"].getint("min_year", fallback=1900)
    max_year = cfg["run"].getint("max_year", fallback=2100)
    similarity_threshold = cfg["run"].getfloat("similarity_threshold", fallback=0.78)

    if not input_dir.exists() or not input_dir.is_dir():
        raise NotADirectoryError(
            f"Carpeta de entrada inválida: {input_dir}. "
            "Crea la carpeta o corrige [input].input_dir en config.ini."
        )

    if args.load_mode == "initial":
        selected_input_dir = input_dir
    elif args.load_mode == "retry":
        selected_input_dir = retry_input_dir
    elif args.non_interactive:
        selected_input_dir = input_dir
    else:
        selected_input_dir = choose_load_mode(input_dir=input_dir, retry_input_dir=retry_input_dir)

    is_retry_mode = selected_input_dir.resolve() == retry_input_dir.resolve()

    if not selected_input_dir.exists() or not selected_input_dir.is_dir():
        raise NotADirectoryError(
            f"Carpeta seleccionada inválida: {selected_input_dir}. "
            "Crea la carpeta o corrige [input].retry_input_dir en config.ini."
        )

    import_actions = dict(cfg["import_actions"]) if "import_actions" in cfg else {}
    action_sql = import_actions.get(table_identifier)

    metadata = get_table_metadata(conn_params, schema, table)
    insert_cols = get_insertable_columns(metadata)

    if action_sql and not is_retry_mode:
        print(f"\n>> [IMPORT_PREVIOUS_ACTION] → acción previa: {action_sql}")

    fixed_cols = list(cfg["fixed_values"].keys()) if "fixed_values" in cfg else []
    target_columns = [c for c in insert_cols if c not in fixed_cols]

    excel_file = pick_excel_file(selected_input_dir, file_name)
    is_xls_source = excel_file.suffix.lower() == ".xls"
    resolved_sheet_name = choose_sheet_name(excel_file, sheet_name)

    header_row_idx = choose_header_mode_interactive(
        cfg=cfg,
        config_path=config_path,
        mapping_path=mapping_path,
        excel_path=excel_file,
        sheet_name=resolved_sheet_name,
        target_columns=target_columns,
        similarity_threshold=similarity_threshold,
        non_interactive=args.non_interactive,
    )

    # 💾 guardar header detectado/usado
    save_header_meta(
        mapping_path,
        table_identifier,
        header_row_idx + 1  # Excel es 1-based
    )

    try:
        excel_display = excel_file.relative_to(input_dir)
    except ValueError:
        excel_display = excel_file.name
    
    print(f"\n>> [INFO] Leyendo Excel: {excel_display} (hoja: {resolved_sheet_name}, header: {header_row_idx + 1})")

    df_excel_raw = read_excel_with_sheet(excel_file, resolved_sheet_name, header_row_idx=header_row_idx)
    df_excel_raw = drop_control_columns(df_excel_raw)
    
    df_raw = df_excel_raw.copy()

    df_raw, _, mapping_df = resolve_mapping(
        df_raw=df_excel_raw.copy(),
        cfg=cfg,
        mapping_path=mapping_path,
        output_dir=output_dir,
        schema=schema,
        table=table,
        target_columns=target_columns,
        similarity_threshold=similarity_threshold,
        auto_approve=args.auto_approve_mapping,
        non_interactive=args.non_interactive,
        auto_confirm_known_mapping=auto_confirm_known_mapping,
        cleanup_mapping_review=cleanup_mapping_review,
    )

    if args.only_mapping:
        print("\n[INFO] Modo --only-mapping: homologación confirmada. No se insertaron datos.")
        return

    df_raw = clean_text_values(df_raw)

    excel_input_cols = [c for c in insert_cols if c not in fixed_cols]
    df_raw = drop_fully_empty_rows(df_raw, [c for c in excel_input_cols if c in df_raw.columns])

    missing_plan = collect_missing_input_columns(
        df=df_raw,
        metadata=metadata,
        insert_cols=insert_cols,
        fixed_cols=fixed_cols,
    )
    
    suggestions_df = print_mapping_coverage_summary(
        df_raw=df_raw,
        mapping_df=mapping_df,
        insert_cols=insert_cols,
        fixed_cols=fixed_cols,
        missing_plan=missing_plan,
        similarity_threshold=similarity_threshold,
    )

    if missing_plan:
        accepted_missing = args.yes_missing_columns or args.non_interactive
    
        if accepted_missing:
            origin = "--yes-missing-columns" if args.yes_missing_columns else "--non-interactive"
            print(f"[VALIDACION] Columnas faltantes aceptadas automáticamente por {origin}.")
        else:
            missing_decision = confirm_missing_columns_plan(
                missing_plan=missing_plan,
                suggestions_df=suggestions_df,
                mapping_path=mapping_path,
                table_section=table_identifier,
                non_interactive=args.non_interactive,
            )
    
            if missing_decision == "reload":
                print("\n>> [INFO] Reaplicando mapping con nueva homologación...")
            
                df_raw, _, mapping_df = resolve_mapping(
                    df_raw=df_excel_raw.copy(),
                    cfg=cfg,
                    mapping_path=mapping_path,
                    output_dir=output_dir,
                    schema=schema,
                    table=table,
                    target_columns=target_columns,
                    similarity_threshold=similarity_threshold,
                    auto_approve=args.auto_approve_mapping,
                    non_interactive=args.non_interactive,
                    auto_confirm_known_mapping=auto_confirm_known_mapping,
                    cleanup_mapping_review=cleanup_mapping_review,
                )
            
                df_raw = clean_text_values(df_raw)
            
                excel_input_cols = [c for c in insert_cols if c not in fixed_cols]
                df_raw = drop_fully_empty_rows(df_raw, [c for c in excel_input_cols if c in df_raw.columns])
            
                missing_plan = collect_missing_input_columns(
                    df=df_raw,
                    metadata=metadata,
                    insert_cols=insert_cols,
                    fixed_cols=fixed_cols,
                )
            
                suggestions_df = print_mapping_coverage_summary(
                    df_raw=df_raw,
                    mapping_df=mapping_df,
                    insert_cols=insert_cols,
                    fixed_cols=fixed_cols,
                    missing_plan=missing_plan,
                    similarity_threshold=similarity_threshold,
                )
            
                if missing_plan:
                    missing_decision = confirm_missing_columns_plan(
                        missing_plan=missing_plan,
                        suggestions_df=suggestions_df,
                        mapping_path=mapping_path,
                        table_section=table_identifier,
                        non_interactive=args.non_interactive,
                    )
            
                    if missing_decision == "cancel":
                        raise UserCancelledLoad(
                            "Carga detenida por usuario antes de insertar datos, por columnas faltantes en el Excel."
                        )
    
            if missing_decision == "cancel":
                raise UserCancelledLoad(
                    "Carga detenida por usuario antes de insertar datos, por columnas faltantes en el Excel."
                )
    
    result = validate_and_transform(
        df_raw=df_raw,
        metadata=metadata,
        insert_cols=insert_cols,
        fixed_cols=fixed_cols,
        min_year=min_year,
        max_year=max_year,
        allow_out_of_range_dates=allow_out_of_range_dates,
    )

    if not result.invalid_df.empty and "errores" in result.invalid_df.columns:
        print("\n>> [DIAGNOSTICO] Top errores de validación:")
        exploded = result.invalid_df["errores"].astype("string").str.split(", ").explode()
        counts = exploded.value_counts(dropna=True).head(10)
        for err, cnt in counts.items():
            print(f"  - {err}: {cnt}")

    valid_df = apply_missing_columns_plan(result.valid_df, missing_plan)
    valid_df = apply_fixed_values(valid_df, cfg, insert_cols)

    invalid_path = export_invalid(result.invalid_df, output_dir)
    
    source_annotated = False
    
    if not result.invalid_df.empty:
        source_annotated = annotate_source_excel_errors(
            excel_path=excel_file,
            sheet_name=resolved_sheet_name,
            error_messages=result.error_messages,
            header_row_idx=header_row_idx,
        )
    
    if source_annotated:
        print("- Excel fuente actualizado con columna 'errores'.")
    elif is_xls_source:
        print("\n>> [INFO] Archivo fuente .xls detectado: se generará versión .xlsx con columna 'errores' al finalizar el proceso.")

    if action_sql and not is_retry_mode:
        success_pre_action = execute_pre_import_action(conn_params, action_sql)
        if not success_pre_action:
            raise RuntimeError(
                f"[{schema}.{table}] No se pudo ejecutar la acción previa. "
                "Revisa permisos/tabla/SQL en config.ini."
            )
            
    inserted = insert_valid_rows(
        df=valid_df,
        conn_params=conn_params,
        schema=schema,
        table=table,
        insert_cols=insert_cols,
        batch_size=batch_size,
        progress_every=progress_every,
    )

    print("\n>> [RESUMEN]")
    print(f"- Filas leídas: {len(df_raw)}")
    print(f"- Filas válidas insertadas: {inserted}")
    print(f"- Filas inválidas: {len(result.invalid_df)}")
    if invalid_path:
        print(f"- Reporte inválidos: {invalid_path}")
    if source_annotated:
        try:
            excel_display = excel_file.relative_to(input_dir)
        except ValueError:
            excel_display = excel_file.name
        
        print(f"- Columna 'errores' actualizada en fuente: {excel_display}")
    elif is_xls_source:
        print("- Fuente .xls detectada: la columna 'errores' se exportará en un nuevo .xlsx.")

    processed_path: Path | None = None
    annotated_done_path: Path | None = None
    
    if len(result.invalid_df) > 0:
        retry_copy = copy_invalid_to_retry(invalid_path, retry_input_dir)
        if retry_copy:
            print(f"- Archivo para reintento generado en: {retry_copy}")
    else:
        retry_copy = None
    
    if inserted > 0 and not is_retry_mode:
        partial = len(result.invalid_df) > 0
        status_suffix = "_PARTIAL_ERROR" if partial else "_OK"
    
        if is_xls_source:
            # 1) mover original .xls como backup
            processed_path = mark_excel_as_processed(
                excel_path=excel_file,
                mode=processed_mode,
                done_dir=processed_dir,
                loaded_suffix=loaded_suffix,
                custom_name_suffix="_BACKUP",
            )
            if processed_path:
                print(f"- Excel fuente .xls movido como respaldo: {processed_path}")
    
            # 2) generar nuevo .xlsx con columna errores
            if processed_mode == "none":
                annotated_base_dir = output_dir
            else:
                annotated_base_dir = processed_dir
    
            annotated_done_path = annotated_base_dir / f"{excel_file.stem}{status_suffix}.xlsx"
            annotated_done_path = export_annotated_xlsx_from_source(
                df_raw=df_raw,
                error_messages=result.error_messages,
                output_path=annotated_done_path,
            )
            print(f"- Excel resultado generado en .xlsx: {annotated_done_path}")
    
        else:
            processed_path = mark_excel_as_processed(
                excel_path=excel_file,
                mode=processed_mode,
                done_dir=processed_dir,
                loaded_suffix=loaded_suffix,
                status_suffix=status_suffix,
            )
            if processed_path:
                estado = "parcial con errores" if partial else "completa"
                try:
                    processed_display = processed_path.relative_to(processed_dir.parent)
                except ValueError:
                    processed_display = processed_path.name
                print(f"- Excel marcado como carga {estado}: {processed_display}")
    
        if partial and retry_copy:
            register_retry_entry(
                index_path=retry_index_path,
                retry_file=retry_copy,
                original_processed_path=annotated_done_path or processed_path,
                invalid_report_path=invalid_path,
            )

    if is_retry_mode and inserted > 0:
        safe_delete(excel_file)
        print("- Excel de reintento procesado y eliminado de inputs_retry.")

    if is_retry_mode and inserted == len(df_raw) and len(result.invalid_df) == 0:
        entry = pop_retry_entry(retry_index_path, excel_file)

        if entry and entry.get("invalid_report_path"):
            safe_delete(Path(entry["invalid_report_path"]))

        if entry and entry.get("original_processed_path"):
            updated = rename_partial_to_ok(Path(entry["original_processed_path"]))
            if updated:
                print(f"- Original parcial actualizado a OK: {updated}")

        print("- Reintento completado: limpiados archivos de retry/salida asociados.")


# =========================================================
# TESTS
# =========================================================


def parse_date_series_unbounded(series: pd.Series) -> pd.Series:
    def parse_one(value):
        if value is None or pd.isna(value):
            return pd.NA

        txt = str(value).strip()
        if not txt:
            return pd.NA

        # serial Excel
        try:
            num = float(txt.replace(",", "."))
        
            # serial Excel válido aproximado hasta 9999-12-31
            if 1 <= num <= 2958465:
                return date(1899, 12, 30) + timedelta(days=int(num))
        
        except Exception:
            pass

        # normalizar si viene con hora: yyyy-mm-dd hh:mm:ss -> yyyy-mm-dd
        if re.fullmatch(r"\d{4}-\d{1,2}-\d{1,2}\s+\d{2}:\d{2}:\d{2}", txt):
            txt = txt.split()[0]

        # dd/mm/yyyy o dd-mm-yyyy
        m = re.fullmatch(r"(\d{1,2})[/-](\d{1,2})[/-](\d{4})", txt)
        if m:
            d, mth, y = map(int, m.groups())
            try:
                return date(y, mth, d)
            except ValueError:
                return pd.NA

        # yyyy-mm-dd
        m = re.fullmatch(r"(\d{4})-(\d{1,2})-(\d{1,2})", txt)
        if m:
            y, mth, d = map(int, m.groups())
            try:
                return date(y, mth, d)
            except ValueError:
                return pd.NA

        return pd.NA

    return series.apply(parse_one)

def run_tests() -> None:
    assert parse_column_default_literal("false")[1] is False
    assert parse_column_default_literal("true")[1] is True
    assert parse_column_default_literal("'abc'::text")[1] == "abc"
    assert parse_column_default_literal("42")[1] == 42
    assert parse_column_default_literal("3.14")[1] == 3.14

    periodo = pd.Series(["ene-24", "dic/2025", "invalido"])
    p = parse_periodo_series(periodo)
    assert pd.notna(p.iloc[0]) and p.iloc[0].month == 1 and p.iloc[0].year == 2024
    assert pd.notna(p.iloc[1]) and p.iloc[1].month == 12 and p.iloc[1].year == 2025
    assert pd.isna(p.iloc[2])

    nums = pd.Series(["1.234,56", "1234,56", "np.float64(12.5)", "x"])
    n = parse_numeric_series(nums)
    assert abs(float(n.iloc[0]) - 1234.56) < 1e-9
    assert abs(float(n.iloc[1]) - 1234.56) < 1e-9
    assert abs(float(n.iloc[2]) - 12.5) < 1e-9
    assert pd.isna(n.iloc[3])

    b = parse_bool_series(pd.Series(["si", "no", "1", "0", "talvez"]))
    assert b.iloc[0] is True
    assert b.iloc[1] is False
    assert b.iloc[2] is True
    assert b.iloc[3] is False
    assert b.iloc[4] == "talvez"

    print("[TEST] OK - pruebas mínimas superadas")


if __name__ == "__main__":
    try:
        main()
    except UserCancelledLoad as e:
        print("\n>> [INFO] Carga detenida por usuario")
        print(f"Detalle: {e}")
        print("Acción: ajusta mapping.ini o la estructura del Excel y vuelve a ejecutar cuando quieras.")
    except FileNotFoundError as e:
        print("\n>> [ERROR] Archivo no encontrado")
        print(f"Detalle: {e}")
        print("Acción: valida rutas en config.ini (input_dir, file_name, config-path) y vuelve a ejecutar.")
    except NotADirectoryError as e:
        print("\n>> [ERROR] Carpeta inválida")
        print(f"Detalle: {e}")
        print("Acción: crea la carpeta indicada o corrige [input].input_dir en config.ini.")
    except KeyError as e:
        print("\n>> [ERROR] Configuración incompleta")
        print(f"Detalle: {e}")
        print("Acción: completa las secciones/campos faltantes en config.ini usando config.example.ini como guía.")
    except ValueError as e:
        print("\n>> [ERROR] Validación/configuración")
        print(f"Detalle: {e}")
        print("Acción: corrige configuración o estructura del Excel según el mensaje y vuelve a ejecutar.")
    except psycopg2.OperationalError as e:
        print("\n>> [ERROR] Conexión a base de datos")
        print(f"Detalle: {e}")
        print("Acción: revisa host/port/dbname/user/password, red/VPN y permisos de acceso.")
    except psycopg2.Error as e:
        print("\n>> [ERROR] Error PostgreSQL")
        print(f"Detalle: {e}")
        print("Acción: revisa schema/tabla/columnas y tipos de datos; ajusta mapping/config y reintenta.")
    except Exception as e:
        print("\n>> [ERROR] Error inesperado")
        print(f"Detalle: {e}")
        print("Acción: revisa el traceback y comparte el error para diagnóstico.")